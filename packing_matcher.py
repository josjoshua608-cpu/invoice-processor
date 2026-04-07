"""
packing_matcher.py
==================
Matches aggregated descriptions against the Packing List sheet to
calculate the total number of packages per HS Code group.

VBA logic replicated exactly:
    - Scan rows 2..lastRow of the packing sheet (column B = description,
      column G = package count)
    - For each description fragment in the "|"-split list, compare
      LCase(packDesc) == LCase(Trim(d))  (case-insensitive exact match)
    - On first match for a given packing row, add the package count and
      break to the next packing row (Exit For)
    - Sum all matched package counts → pkgTotal

Column mapping (packing sheet, 1-based):
    B = column 2  → description
    G = column 7  → number of packages
"""

from __future__ import annotations

import logging
from typing import Dict, List

from openpyxl.worksheet.worksheet import Worksheet

from aggregator import AggregatedRow

logger = logging.getLogger(__name__)

PACK_COL_DESC = 2   # Column B
PACK_COL_PKG  = 7   # Column G
PACK_START_ROW = 2  # Row 1 assumed to be headers


def _cell_str(ws: Worksheet, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(str(val).strip().replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def build_package_totals(
    pack_ws: Worksheet,
    aggregated: Dict[str, AggregatedRow],
) -> Dict[str, float]:
    """
    Calculate package totals for each HS Code by matching descriptions
    against the Packing List sheet.

    Args:
        pack_ws:    Packing List worksheet.
        aggregated: Dict of {hs_code: AggregatedRow} from aggregator.

    Returns:
        Dict {hs_code: total_packages} for every HS Code in aggregated.
    """
    # Detect last meaningful row in packing sheet (column B)
    pack_last_row = pack_ws.max_row or 1000

    # Pre-load packing rows into memory for O(1) repeated access
    # (packing sheets are small; this avoids repeated cell reads)
    packing_rows: List[tuple[str, float]] = []
    for r in range(PACK_START_ROW, pack_last_row + 1):
        desc = _cell_str(pack_ws, r, PACK_COL_DESC)
        pkg_count = _safe_float(pack_ws.cell(row=r, column=PACK_COL_PKG).value)
        if desc:
            packing_rows.append((desc, pkg_count))

    logger.info(
        "Packing list loaded: %d non-empty description rows.", len(packing_rows)
    )

    pkg_totals: Dict[str, float] = {}

    for hs_code, agg_row in aggregated.items():
        pkg_total = 0.0
        desc_list: List[str] = agg_row.descriptions  # already split

        for pack_desc, pack_count in packing_rows:
            pack_desc_lower = pack_desc.lower()

            matched = False
            for d in desc_list:
                if pack_desc_lower == d.strip().lower():
                    pkg_total += pack_count
                    matched = True
                    break  # mirrors VBA "Exit For" after first match per row

            if matched:
                logger.debug(
                    "HS %s: packing row '%s' matched → +%.0f pkgs",
                    hs_code,
                    pack_desc,
                    pack_count,
                )

        pkg_totals[hs_code] = pkg_total
        logger.debug(
            "HS %s total packages = %.0f", hs_code, pkg_total
        )

    return pkg_totals
