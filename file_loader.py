"""
file_loader.py
==============
Loads an Excel invoice workbook and detects the Invoice sheet and the
Packing List sheet by name (case-insensitive "Invoice" match) or by
sheet index (sheet 2 as fallback for packing list — matching VBA logic).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class FileLoaderError(Exception):
    """Raised when the workbook cannot be loaded or required sheets are missing."""


def load_workbook_safe(file_path: str | Path) -> openpyxl.Workbook:
    """
    Load an Excel workbook in read-only mode.

    Args:
        file_path: Absolute or relative path to the .xlsx file.

    Returns:
        openpyxl.Workbook instance.

    Raises:
        FileLoaderError: If the file does not exist or cannot be opened.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileLoaderError(f"File not found: {path}")
    if not path.suffix.lower() in {".xlsx", ".xlsm"}:
        raise FileLoaderError(
            f"Unsupported file type '{path.suffix}'. Expected .xlsx or .xlsm."
        )

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        logger.info("Loaded workbook: %s  |  sheets: %s", path.name, wb.sheetnames)
        return wb
    except Exception as exc:
        raise FileLoaderError(f"Cannot open workbook '{path}': {exc}") from exc


def detect_sheets(wb: openpyxl.Workbook) -> Tuple[Worksheet, Worksheet]:
    """
    Detect the Invoice sheet and the Packing List sheet.

    Invoice sheet: first sheet whose name contains "invoice" (case-insensitive).
    Packing list:  the *second* sheet in the workbook (VBA: wbSource.Sheets(2)),
                   unless a sheet name contains "pack" (case-insensitive), in
                   which case the named match takes priority.

    Args:
        wb: Loaded openpyxl.Workbook.

    Returns:
        (invoice_sheet, packing_sheet) tuple of Worksheet objects.

    Raises:
        FileLoaderError: If the Invoice sheet cannot be found.
    """
    sheet_names = wb.sheetnames

    # --- Invoice sheet ---
    invoice_ws: Worksheet | None = None
    for name in sheet_names:
        if "invoice" in name.lower():
            invoice_ws = wb[name]
            logger.info("Invoice sheet detected: '%s'", name)
            break

    if invoice_ws is None:
        raise FileLoaderError(
            f"Invoice sheet not found. Available sheets: {sheet_names}"
        )

    # --- Packing List sheet ---
    # Priority 1: sheet whose name contains "pack"
    packing_ws: Worksheet | None = None
    for name in sheet_names:
        if "pack" in name.lower():
            packing_ws = wb[name]
            logger.info("Packing List sheet detected by name: '%s'", name)
            break

    # Priority 2: sheet index 1 (0-based → second sheet), mirroring VBA Sheets(2)
    if packing_ws is None:
        if len(sheet_names) >= 2:
            packing_ws = wb[sheet_names[1]]
            logger.info(
                "Packing List sheet inferred from position 2: '%s'", sheet_names[1]
            )
        else:
            # Tolerate single-sheet workbooks (packing data missing)
            packing_ws = invoice_ws
            logger.warning(
                "Only one sheet found; packing list defaults to Invoice sheet."
            )

    return invoice_ws, packing_ws
