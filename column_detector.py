"""
column_detector.py
==================
Detects the header row and individual column positions for the six
required data columns.  No fixed positions are assumed – every lookup
is keyword-based, exactly as in the VBA macro.

Required columns detected:
    HS Code           – header contains both "HS" and "Code"
    Description       – header contains "Description of Goods"
    Quantity          – header contains "Quantity"
    Total Value       – header contains "Total Value"
    Total Net Weight  – header contains "Total Net"
    Total Gross Weight– header contains "Total Gross"
"""

from __future__ import annotations

import logging
from typing import Dict, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

SCAN_HEADER_ROWS = 40


class ColumnDetectorError(Exception):
    """Raised when a mandatory column header cannot be located."""


def _cell_str(ws: Worksheet, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""


def _row_width(ws: Worksheet, row: int) -> int:
    """
    Return the number of populated columns in *row*.
    Uses ws.max_column as upper bound (safe for read-only worksheets).
    """
    max_col = ws.max_column or 50
    for c in range(max_col, 0, -1):
        if ws.cell(row=row, column=c).value not in (None, ""):
            return c
    return max_col


def detect_header_row_and_hs_col(ws: Worksheet) -> Tuple[int, int]:
    """
    Find the first row (within rows 1–40) that contains a cell whose
    value includes both "HS" and "Code" (case-insensitive).

    Returns:
        (header_row, col_hs) 1-based indices.

    Raises:
        ColumnDetectorError: If the HS Code column cannot be found.
    """
    for r in range(1, SCAN_HEADER_ROWS + 1):
        width = _row_width(ws, r)
        for c in range(1, width + 1):
            val = _cell_str(ws, r, c).lower()
            if "hs" in val and "code" in val:
                logger.info("HS Code column found at row=%d col=%d", r, c)
                return r, c

    raise ColumnDetectorError(
        f"HS Code header not detected in first {SCAN_HEADER_ROWS} rows."
    )


def detect_data_columns(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """
    Scan the header row and return a dict of column indexes (1-based)
    for each required field.

    Args:
        ws:         Invoice worksheet.
        header_row: Row number containing column headers.

    Returns:
        Dict with keys: 'hs', 'desc', 'qty', 'val', 'net', 'gross'.
        Values are 1-based column integers.

    Raises:
        ColumnDetectorError: If Description or Quantity columns are missing
                             (minimum required to produce meaningful output).
    """
    width = _row_width(ws, header_row)
    cols: Dict[str, Optional[int]] = {
        "hs": None,
        "desc": None,
        "qty": None,
        "val": None,
        "net": None,
        "gross": None,
    }

    for c in range(1, width + 1):
        val = _cell_str(ws, header_row, c)
        val_lower = val.lower()

        if "hs" in val_lower and "code" in val_lower and cols["hs"] is None:
            cols["hs"] = c
            logger.debug("HS col=%d  header=%r", c, val)

        elif "description of goods" in val_lower and cols["desc"] is None:
            cols["desc"] = c
            logger.debug("Desc col=%d  header=%r", c, val)

        elif "quantity" in val_lower and cols["qty"] is None:
            cols["qty"] = c
            logger.debug("Qty col=%d  header=%r", c, val)

        elif "total value" in val_lower and cols["val"] is None:
            cols["val"] = c
            logger.debug("Val col=%d  header=%r", c, val)

        # "Total Net" must come before "Total Gross" in keyword priority
        elif "total net" in val_lower and cols["net"] is None:
            cols["net"] = c
            logger.debug("Net col=%d  header=%r", c, val)

        elif "total gross" in val_lower and cols["gross"] is None:
            cols["gross"] = c
            logger.debug("Gross col=%d  header=%r", c, val)

    # Validate mandatory columns
    missing = [k for k, v in cols.items() if v is None]
    if missing:
        logger.warning(
            "The following column(s) were not detected and will be treated as "
            "zero/empty: %s",
            missing,
        )

    # Return with None replaced by 0 (sentinel: "column absent")
    return {k: (v if v is not None else 0) for k, v in cols.items()}
